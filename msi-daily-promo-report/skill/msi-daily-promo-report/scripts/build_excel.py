# -*- coding: utf-8 -*-
"""
MSI 每日活動報告產表腳本
用法: python build_excel.py <input.json> <output.xlsx>
輸入 JSON 結構見 SKILL.md 步驟 3。
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CELL = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=13, color="1F3864")
NOTE = Font(name=FONT, size=9, italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws, row, ncols):
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, WRAPC, BORDER
    ws.row_dimensions[row].height = 28


def safe_sheet_name(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:31]


def main(src, dst):
    with open(src, encoding="utf-8") as f:
        data = json.load(f)
    crawl = data.get("crawl_time", "")
    cols = data.get("collections", [])
    promos = data.get("promotions", [])

    wb = Workbook()

    # ---------- 摘要頁 ----------
    ws = wb.active
    ws.title = "00_摘要"
    ws["A1"] = "MSI Store 每日活動優惠報告"
    ws["A1"].font = Font(name=FONT, bold=True, size=15, color="1F3864")
    ws["A2"] = f"資料抓取時間：{crawl}"
    ws["A2"].font = NOTE
    headers = ["活動專區", "商品數", "有庫存", "有折扣品項", "最深折扣", "折扣方式", "備註"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers))
    r = 5
    for c in cols:
        items = c.get("items", [])
        disc = [i for i in items if i.get("compare_at")
                and i["compare_at"] > (i.get("price") or 0)]
        deepest = (max(1 - i["price"] / i["compare_at"] for i in disc)
                   if disc else None)
        row = [c["name"], len(items),
               sum(1 for i in items if i.get("available")),
               len(disc),
               f"{deepest:.0%}" if deepest is not None else "—",
               c.get("mech", ""), c.get("note", "")]
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font, cell.alignment, cell.border = CELL, WRAP, BORDER
        r += 1
    if promos:
        r += 1
        ws.cell(row=r, column=1,
                value="【登錄送／贈品類活動（折扣方式 C：需至活動頁登錄，贈品鑑賞期後寄送）】"
                ).font = Font(name=FONT, bold=True, size=10)
        r += 1
        ph = ["活動名稱", "活動期間", "贈品內容", "狀態"]
        for i, h in enumerate(ph, 1):
            ws.cell(row=r, column=i, value=h)
        style_header(ws, r, len(ph))
        r += 1
        for p in promos:
            row = [p.get("title", ""), p.get("period", ""),
                   p.get("gift", ""), p.get("status", "")]
            for i, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.font, cell.alignment, cell.border = CELL, WRAP, BORDER
            r += 1
    for i, w in enumerate([24, 9, 9, 11, 10, 52, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- 各專區分頁 ----------
    for c in cols:
        sh = wb.create_sheet(safe_sheet_name(c["name"]))
        sh["A1"] = f"{c['name']}｜折扣方式：{c.get('mech', '')}"
        sh["A1"].font = TITLE
        sh["A2"] = f"抓取時間：{crawl}" + (f"　※{c['note']}" if c.get("note") else "")
        sh["A2"].font = NOTE
        headers = ["序號", "商品名稱", "分類", "現售價 NT$", "原價 NT$",
                   "折抵 NT$", "折扣率", "庫存", "連結"]
        for i, h in enumerate(headers, 1):
            sh.cell(row=4, column=i, value=h)
        style_header(sh, 4, len(headers))
        for idx, it in enumerate(c.get("items", []), 1):
            r = 4 + idx
            has_disc = bool(it.get("compare_at")
                            and it["compare_at"] > (it.get("price") or 0))
            row = [idx, it.get("title", ""), it.get("type", ""),
                   it.get("price"),
                   it.get("compare_at") if has_disc else "—",
                   f"=E{r}-D{r}" if has_disc else "—",
                   f"=1-D{r}/E{r}" if has_disc else "—",
                   "有庫存" if it.get("available") else "售完/缺貨",
                   it.get("url", "")]
            for i, v in enumerate(row, 1):
                cell = sh.cell(row=r, column=i, value=v)
                cell.font, cell.alignment, cell.border = CELL, WRAP, BORDER
            for col in "DEF":
                sh[f"{col}{r}"].number_format = "#,##0"
            sh[f"G{r}"].number_format = "0%"
        for i, w in enumerate([5, 52, 16, 12, 12, 11, 9, 12, 50], 1):
            sh.column_dimensions[get_column_letter(i)].width = w
        sh.freeze_panes = "A5"
        n = len(c.get("items", []))
        if n:
            sh.auto_filter.ref = f"A4:I{4 + n}"

    wb.save(dst)
    print(f"saved {dst}: {sum(len(c.get('items', [])) for c in cols)} items "
          f"across {len(cols)} collections")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("usage: build_excel.py <input.json> <output.xlsx>")
    main(sys.argv[1], sys.argv[2])
